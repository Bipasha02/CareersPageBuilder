# server/app/main.py
from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Body, Request
from fastapi.middleware.cors import CORSMiddleware
# from pydantic import BaseModel
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session
from typing import Optional
from pathlib import Path
from uuid import uuid4
from datetime import datetime
import json
from . import models, database
from .database import engine, SessionLocal, Base
from .models import Company, Job, Section
from .database import init_db, get_all_jobs, import_seed_data

# -----------------------
# App Setup
# -----------------------
app = FastAPI(title="Careers Page Builder API")

# Create tables on startup
models.Base.metadata.create_all(bind=database.engine)

# -----------------------
# CORS Middleware
# -----------------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # replace with your frontend origin if needed
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------
# Include Custom Routes
# -----------------------
from app.routes import feature_settings
app.include_router(feature_settings.router)

# -----------------------
# Root Route 
# -----------------------
@app.get("/")
def read_root():
    return {"message": "FastAPI backend running successfully"}

# -----------------------
# Utility
# -----------------------
def slugify(s: Optional[str], unique_id: Optional[str] = None) -> str:
    s = s or ""
    base = "".join(ch.lower() if ch.isalnum() else "-" for ch in s).strip("-") or "item"
    return f"{base}-{unique_id[:6]}" if unique_id else base


# -----------------------
# DB Dependency
# -----------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(bind=engine)


# -----------------------
# Upload XLSX → seed.json
# -----------------------
@app.post("/api/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...)):
    """Accepts an Excel (.xlsx) file and converts it into a structured seed.json file."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    from openpyxl import load_workbook

    BASE_DIR = Path(__file__).resolve().parent.parent
    DATA_PATH = BASE_DIR / "data" / "seed.json"

    content = await file.read()
    temp_path = DATA_PATH.parent / "uploaded.xlsx"
    temp_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path.write_bytes(content)

    wb = load_workbook(temp_path)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(headers, r)) for r in sheet.iter_rows(min_row=2, values_only=True)]

    company_slug = slugify("Imported Company", "co1")
    companies = [{
        "id": "co_1",
        "slug": company_slug,
        "name": "Imported Company",
        "is_published": True
    }]

    sections = [{
        "id": "sec_1",
        "company_id": "co_1",
        "type": "about",
        "title": "About Imported Company",
        "content": "Imported automatically from Excel upload.",
        "position": 1,
        "is_visible": True
    }]

    jobs = []
    for idx, r in enumerate(rows, start=1):
        title = r.get("title") or r.get("job title") or f"Role {idx}"
        slug = slugify(title, str(uuid4()))
        jobs.append({
            "id": f"job_{idx}",
            "company_id": "co_1",
            "title": title,
            "slug": slug,
            "location": r.get("location") or "",
            "type": r.get("employment_type") or r.get("job_type") or "",
            "description": r.get("description") or "",
            "published": True,
            "apply_url": r.get("apply_url") or "",
        })

    seed = {"companies": companies, "sections": sections, "jobs": jobs}
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(seed, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"ok": True, "imported": len(jobs)}


# -----------------------
# Import seed.json → DB
# -----------------------
@app.post("/api/import-seed")
def import_seed(db: Session = Depends(get_db)):
    BASE_DIR = Path(__file__).resolve().parent.parent
    DATA_PATH = BASE_DIR / "data" / "seed.json"

    if not DATA_PATH.exists():
        raise HTTPException(status_code=404, detail="seed.json not found in /data")

    raw = json.loads(DATA_PATH.read_text(encoding="utf-8"))

    try:
        # 1️⃣ Clear previous data
        from sqlalchemy import text
        with db.begin():
            db.execute(text("DELETE FROM job"))
            db.execute(text("DELETE FROM section"))
            db.execute(text("DELETE FROM company"))

        # 2️⃣ Add all companies first
        company_map = {}
        for idx, c in enumerate(raw.get("companies", []), start=1):
            cid = c.get("id") or f"co_{idx}"
            c_slug = c.get("slug") or slugify(c.get("name"), cid)
            comp = Company(
                id=str(cid),
                slug=c_slug,
                name=c.get("name"),
                website=c.get("website"),
                logo=c.get("logo"),
                banner_url=c.get("banner_url"),
                video_url=c.get("video_url"),
                description=c.get("description"),
                theme_color=c.get("theme_color") or "#0a66c2",
            )
            db.add(comp)  # Use add(), not merge()
            company_map[c_slug] = cid

        db.commit()  # Commit all companies together

        # 3️⃣ Add sections
        for idx, s in enumerate(raw.get("sections", []), start=1):
            sec = Section(
                id=s.get("id") or f"sec_{idx}",
                company_id=company_map.get(s.get("company_id"), s.get("company_id")),
                type=s.get("type"),
                title=s.get("title"),
                content=s.get("content") or "",
                position=s.get("position") or 1,
                is_visible=bool(s.get("is_visible", True)),
            )
            db.add(sec)

        # 4️⃣ Add jobs
        for idx, j in enumerate(raw.get("jobs", []), start=1):
            job = Job(
                id=j.get("id") or f"job_{idx}",
                company_id=company_map.get(j.get("company_id"), j.get("company_id")),
                title=j.get("title"),
                slug=slugify(j.get("title"), str(uuid4())),
                location=j.get("location"),
                type=j.get("type"),
                description=j.get("description"),
                published=bool(j.get("published", True)),
                apply_url=j.get("apply_url"),
            )
            db.add(job)

        db.commit()  # Commit sections and jobs together
        return {"ok": True, "message": "Seed imported successfully"}
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))



# -----------------------
# Public APIs
# -----------------------
@app.get("/api/companies")
def list_companies(db: Session = Depends(get_db)):
    return jsonable_encoder(db.query(Company).all())


@app.get("/api/companies/{slug}")
def get_company_public(slug: str, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.slug == slug).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    sections = db.query(Section).filter(Section.company_id == company.id).order_by(Section.position).all()
    jobs = db.query(Job).filter(Job.company_id == company.id, Job.published == True).all()
    return {"company": company, "sections": sections, "jobs": jobs}


@app.get("/api/jobs")
def list_jobs(
    q: Optional[str] = None,
    location: Optional[str] = None,
    type: Optional[str] = None,
    company: Optional[str] = None,
    db: Session = Depends(get_db)
):
    stmt = db.query(Job)
    if q:
        stmt = stmt.filter(Job.title.ilike(f"%{q}%"))
    if location:
        stmt = stmt.filter(Job.location == location)
    if type:
        stmt = stmt.filter(Job.type == type)
    if company:
        stmt = stmt.filter(Job.company_id == company)
    return jsonable_encoder(stmt.all())


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return jsonable_encoder(job)


# -----------------------
# Admin APIs
# -----------------------
@app.put("/api/company/{company_id}/update")
def update_company(company_id: str, data: dict = Body(...), db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    allowed_fields = ["name", "logo", "banner_url", "video_url", "theme_color", "description"]
    for key, value in data.items():
        if key in allowed_fields and value is not None:
            setattr(company, key, value)

    db.commit()
    db.refresh(company)
    return {"ok": True, "company": company}


@app.post("/api/company/{company_id}/sections")
def add_section(company_id: str, section: dict = Body(...), db: Session = Depends(get_db)):
    new_section = Section(
        id=str(uuid4()),
        company_id=company_id,
        type=section.get("type"),
        title=section.get("title"),
        content=section.get("content"),
        position=section.get("position", 1),
        is_visible=section.get("is_visible", True),
    )
    db.add(new_section)
    db.commit()
    db.refresh(new_section)
    return {"ok": True, "section": new_section}


@app.put("/api/section/{section_id}")
def update_section(section_id: str, data: dict = Body(...), db: Session = Depends(get_db)):
    section = db.query(Section).filter(Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    for key, value in data.items():
        if hasattr(section, key) and value is not None:
            setattr(section, key, value)

    db.commit()
    db.refresh(section)
    return {"ok": True, "section": section}


@app.delete("/api/section/{section_id}")
def delete_section(section_id: str, db: Session = Depends(get_db)):
    section = db.query(Section).filter(Section.id == section_id).first()
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")

    db.delete(section)
    db.commit()
    return {"ok": True, "deleted": section_id}


@app.get("/api/company/{company_id}/preview")
def preview_company_page(company_id: str, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    sections = db.query(Section).filter(
        Section.company_id == company.id, Section.is_visible == True
    ).order_by(Section.position).all()

    jobs = db.query(Job).filter(
        Job.company_id == company.id, Job.published == True
    ).all()

    return {"company": company, "sections": sections, "jobs": jobs}

# ______________

@app.post("/api/save-features")
async def save_features(request: Request):
    try:
        data = await request.json()
        print("✅ Received raw data:", data)
    except Exception as e:
        print(".", e)
        data = {}

    return {"ok": True, "message": "Features saved", "data": data}


# -----------------------
# Health Check
# -----------------------
@app.get("/health")
def health_check():
    return {"status": "ok", "time": datetime.utcnow()}
